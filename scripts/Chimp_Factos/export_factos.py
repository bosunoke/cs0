#########################################################################################
import os
import time
import csv
from openpyxl import *
from openpyxl.styles import PatternFill

# Init custom variable
xlsxfilename = 'extract_factos.xlsx'

#########################################################################################

# Get current path
crpth = (os.getcwd())

# Get current date
currentdate = time.strftime('%Y%m%d')

# Create outpout folder for csv
outpoutpathcsv = crpth + '/outpout_csv/' + currentdate
if not os.path.isdir(outpoutpathcsv):
    os.mkdir(outpoutpathcsv)

# Create outpout folder for mailchimp
outpoutpathmcp = crpth + '/outpout_chimp/' + currentdate
if not os.path.isdir(outpoutpathmcp):
    os.mkdir(outpoutpathmcp)

# Create outpout folder for ics
outpoutpathics = crpth + '/outpout_vcf/' + currentdate
if not os.path.isdir(outpoutpathics):
    os.mkdir(outpoutpathics)

# Create outpout folder for error factos
outpoutpathfct = crpth + '/error_factos/' + currentdate
if not os.path.isdir(outpoutpathfct):
    os.mkdir(outpoutpathfct)

# Variables init
lststudent = {}
lstprimary = []
lstmaternelle = []
lstelementaire = []
lstsecondary = []
lstcollege = []
lstlycee = []
lstparents = []
lstclasses = []
lstbademails = []

# Constant
matfstlt = ['P','M','G']
elemfstlt = ['C']
clgefstlt = ['6','5','4','3']
lyceefstlt = ['2','1','T']

#########################################################################################

# Read csv file for bad emails from MailChimp
cvsbepath = crpth + '/bad_emails/bad_emails_mailchimp.csv'
with open(cvsbepath) as csvbe:
    reader = csv.DictReader(csvbe)
    for row in reader:
        lstbademails.append(row['Email Address'])
csvbe.close()

# Access path to the xlsx source file with 10 columns in this specific order :
# Prénom élève, Nom de l'élève, Classe, Autre matricule
# Prénom du responsable 1, Nom du responsable 1, Email Resp1
# Prénom du responsable 2, Nom du responsable 2, Email Resp2,
# Opening the xlsx workbook
os.chdir("/c/src0/cs0/scripts/chimp_factos") #added
wb = load_workbook(filename=crpth + '/' + xlsxfilename, data_only=True, read_only=True)
ws = wb['Export']

print("\nExcel file reading in progress\n")

indexrow = 1
indextuple = 1

# Read the xlsx source file and store information in dict/tuple
for row in ws.rows:
    if indexrow > 1:
        lststudent['fnstudent', indextuple] = ws["{}{}".format("A", indexrow)].value
        lststudent['lnstudent', indextuple] = ws["{}{}".format("B", indexrow)].value
        lststudent['classname', indextuple] = ws["{}{}".format("C", indexrow)].value
        lststudent['matricule', indextuple] = ws["{}{}".format("D", indexrow)].value
        lststudent['fnresp1', indextuple] = ws["{}{}".format("E", indexrow)].value
        lststudent['lnresp1', indextuple] = ws["{}{}".format("F", indexrow)].value
        lststudent['emailresp1', indextuple] = ws["{}{}".format("G", indexrow)].value
        lststudent['fnresp2', indextuple] = ws["{}{}".format("H", indexrow)].value
        lststudent['lnresp2', indextuple] = ws["{}{}".format("I", indexrow)].value
        lststudent['emailresp2', indextuple] = ws["{}{}".format("J", indexrow)].value

        # Populate list of classes
        if lststudent['classname', indextuple] and lststudent['classname', indextuple] not in lstclasses:
            lstclasses.append(lststudent['classname', indextuple])
        # Populate list of unique emails for parents of secondary without bad en empty
        if lststudent['emailresp1', indextuple] and len(lststudent['emailresp1', indextuple]) > 0 and \
                        lststudent['emailresp1', indextuple] not in lstbademails:
            if lststudent['classname', indextuple][0].isnumeric() or lststudent['classname', indextuple][0] == "T":
                if lststudent['emailresp1', indextuple] not in lstsecondary:
                    lstsecondary.append(lststudent['emailresp1', indextuple])
                    if lststudent['classname', indextuple][0] in clgefstlt:
                        if lststudent['emailresp1', indextuple] not in lstcollege:
                            lstcollege.append(lststudent['emailresp1', indextuple])
                    if lststudent['classname', indextuple][0] in lyceefstlt:
                        if lststudent['emailresp1', indextuple] not in lstlycee:
                            lstlycee.append(lststudent['emailresp1', indextuple])
            else:
                if lststudent['emailresp1', indextuple] not in lstprimary:
                    lstprimary.append(lststudent['emailresp1', indextuple])
                    if lststudent['classname', indextuple][0] in matfstlt:
                        if lststudent['emailresp1', indextuple] not in lstmaternelle:
                            lstmaternelle.append(lststudent['emailresp1', indextuple])
                    if lststudent['classname', indextuple][0] in elemfstlt:
                        if lststudent['emailresp1', indextuple] not in lstelementaire:
                            lstelementaire.append(lststudent['emailresp1', indextuple])
        # Populate list of unique emails for parents of primary without bad en empty
        if lststudent['emailresp2', indextuple] and len(lststudent['emailresp2', indextuple]) > 0 and \
                        lststudent['emailresp2', indextuple] not in lstbademails:
            if lststudent['classname', indextuple][0].isnumeric() or lststudent['classname', indextuple][
                0] == "T":
                if lststudent['emailresp2', indextuple] not in lstsecondary:
                    lstsecondary.append(lststudent['emailresp2', indextuple])
                    if lststudent['classname', indextuple][0] in clgefstlt:
                        if lststudent['emailresp2', indextuple] not in lstcollege:
                            lstcollege.append(lststudent['emailresp2', indextuple])
                    if lststudent['classname', indextuple][0] in lyceefstlt:
                        if lststudent['emailresp2', indextuple] not in lstlycee:
                            lstlycee.append(lststudent['emailresp2', indextuple])
            else:
                if lststudent['emailresp2', indextuple] not in lstprimary:
                    lstprimary.append(lststudent['emailresp2', indextuple])
                    if lststudent['classname', indextuple][0] in matfstlt:
                        if lststudent['emailresp2', indextuple] not in lstmaternelle:
                            lstmaternelle.append(lststudent['emailresp2', indextuple])
                    if lststudent['classname', indextuple][0] in elemfstlt:
                        if lststudent['emailresp2', indextuple] not in lstelementaire:
                            lstelementaire.append(lststudent['emailresp2', indextuple])
        indextuple += 1
    indexrow += 1

#########################################################################################

# Create Excel file error Factos
wbfct = Workbook()
xslxfct = outpoutpathfct + '/' + 'erreurs_factos.xlsx'
if os.path.isfile(xslxfct):
    os.remove(xslxfct)
ws1 = wbfct.active
ws1.title = "Erreurs"
ws1['A1'] = "Prénom élève"
ws1['B1'] = "Nom élève"
ws1['C1'] = "Classe"
ws1['D1'] = "Email resp1"
ws1['E1'] = "Email resp2"
indexwbfct = 1

# Create csv file for parents all parents in MailChimp
pathmailchimp = outpoutpathmcp + "/parents_lflp_mailchimp.csv"
csvmailchimp = open(pathmailchimp, "w")
csvmailchimp.write('"firstname","lastname","email","primary","maternelle","elementaire","secondary","college","lycee"\n')

# Create csv file for each class
for entry in lstclasses:
    pathfilecsvadm = crpth + '/outpout_csv/' + currentdate + '/' + str(entry) + '.csv'
    if os.path.isfile(pathfilecsvadm):
        os.remove(pathfilecsvadm)
    csvadm = open(pathfilecsvadm,"w",encoding="cp1252")
    csvadm.write('"firstname","lastname","email"\n')
    csvadm.close()
# Create vcf file for each class
for entry in lstclasses:
    pathfilevcftea = crpth + '/outpout_vcf/' + currentdate + '/' + str(entry) + '.vcf'
    if os.path.isfile(pathfilecsvadm):
        os.remove(pathfilecsvadm)
    csvvcf = open(pathfilevcftea,"w")
    csvvcf.close()

#########################################################################################

# Read the tuple Dict for writing files
indextd = 1
while indextd < indextuple:
    badresult1 = 0
    badresult2 = 0
    if lststudent['fnstudent', indextd] is not None:
        # Open csv file for classes
        pathfilecsvadm = crpth + '/outpout_csv/' + currentdate + '/' + lststudent['classname', indextd] + '.csv'
        csvadm = open(pathfilecsvadm, "a",encoding="cp1252")
        # Open vcf file for classes
        pathfilevcftea = crpth + '/outpout_vcf/' + currentdate + '/' + lststudent['classname', indextd] + '.vcf'
        csvvcf = open(pathfilevcftea, "a")
        # If email 1 is good
        if lststudent['emailresp1', indextd] and len(lststudent['emailresp1', indextd]) > 0 and \
                        lststudent['emailresp1', indextd] not in lstbademails:
            # Writing csv file for resp1
            csvadm.write('"' + lststudent['fnstudent', indextd] + " " + lststudent['lnstudent', indextd] \
                               + '","' + "- " + lststudent['matricule', indextd] + " Resp1" + '","' +  \
                               lststudent['emailresp1', indextd] + '"\n')
            # Writing vcf file for resp1
            csvvcf.write('BEGIN:VCARD\nVERSION:3.0\nN:' + lststudent['fnstudent', indextd] + " " +  \
                         lststudent['lnstudent', indextd] + ';' + "- " + lststudent['matricule', indextd] + \
                         " Resp1" + ';;;\nFN:' + lststudent['fnstudent', indextd] + " " + \
                         lststudent['lnstudent', indextd] + " " + "- " + lststudent['matricule', indextd]  \
                         + " Resp1" + '\nEMAIL;TYPE=INTERNET;TYPE=WORK:' + lststudent['emailresp1', indextd] +  \
                            '\nCATEGORIES:Classe ' + lststudent['classname', indextd] + ' LFLP\nEND:VCARD\n')
            # If parent is not already listed
            if lststudent['emailresp1', indextd] not in lstparents:
                lstparents.append(lststudent['emailresp1', indextd])
                csvmailchimp.write('"' + lststudent['fnresp1', indextd] + '","' + lststudent['lnresp1', indextd] \
                        + '","' + lststudent['emailresp1', indextd] + '",')
                if lststudent['emailresp1', indextd] in lstprimary :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp1', indextd] in lstmaternelle :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp1', indextd] in lstelementaire :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp1', indextd] in lstsecondary :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp1', indextd] in lstcollege :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp1', indextd] in lstlycee :
                    csvmailchimp.write('"1"\n')
                else:
                    csvmailchimp.write('"0"\n')
        else:
            badresult1 = 1
        if lststudent['emailresp2', indextd] and len(lststudent['emailresp2', indextd]) > 0 and \
                        lststudent['emailresp2', indextd] not in lstbademails:
            # Writing csv file for resp1
            csvadm.write('"' + lststudent['fnstudent', indextd] + " " + lststudent['lnstudent', indextd] \
                               + '","' + "- " + lststudent['matricule', indextd] + " Resp2" + '","' +  \
                               lststudent['emailresp2', indextd] + '"\n')
            # Writing vcf file for resp1
            csvvcf.write('BEGIN:VCARD\nVERSION:3.0\nN:' + lststudent['fnstudent', indextd] + " " +  \
                         lststudent['lnstudent', indextd] + ';' + "- " + lststudent['matricule', indextd] + \
                         " Resp2" + ';;;\nFN:' + lststudent['fnstudent', indextd] + " " + \
                         lststudent['lnstudent', indextd] + " " + "- " + lststudent['matricule', indextd]  \
                         + " Resp2" + '\nEMAIL;TYPE=INTERNET;TYPE=WORK:' + lststudent['emailresp2', indextd] +  \
                            '\nCATEGORIES:Classe ' + lststudent['classname', indextd] + ' LFLP\nEND:VCARD\n')
            # If parent is not already listed
            if lststudent['emailresp2', indextd] not in lstparents:
                lstparents.append(lststudent['emailresp2', indextd])
                csvmailchimp.write('"' + lststudent['fnresp2', indextd] + '","' + lststudent['lnresp2', indextd] \
                        + '","' + lststudent['emailresp2', indextd] + '",')
                if lststudent['emailresp2', indextd] in lstprimary :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp2', indextd] in lstmaternelle :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp2', indextd] in lstelementaire :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp2', indextd] in lstsecondary :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp2', indextd] in lstcollege :
                    csvmailchimp.write('"1",')
                else:
                    csvmailchimp.write('"0",')
                if lststudent['emailresp2', indextd] in lstlycee :
                    csvmailchimp.write('"1"\n')
                else:
                    csvmailchimp.write('"0"\n')
        else:
            badresult2 = 1
        # If errors reporting in Excel file
        if badresult1 == 1 or badresult2 == 1:
            indexwbfct += 1
            if badresult1 == 1:
                ws1.cell(row=indexwbfct, column=4).fill = PatternFill(fill_type = "solid", start_color='EE1111', \
                                                                      end_color='EE1111')
            if badresult2 == 1:
                ws1.cell(row=indexwbfct, column=5).fill = PatternFill(fill_type = "solid", start_color='EE1111', \
                                                                      end_color='EE1111')
            ws1.cell(row=indexwbfct, column=1).value = lststudent['fnstudent', indextd]
            ws1.cell(row=indexwbfct, column=2).value = lststudent['lnstudent', indextd]
            ws1.cell(row=indexwbfct, column=3).value = lststudent['classname', indextd]
            ws1.cell(row=indexwbfct, column=4).value = lststudent['emailresp1', indextd]
            ws1.cell(row=indexwbfct, column=5).value = lststudent['emailresp2', indextd]
        # Closing files
        csvvcf.close()
        csvadm.close()
        print("Export of student number " + str(indextd) + " Done !")
    indextd += 1

csvmailchimp.close()
wbfct.save(filename = xslxfct)
nblines = indextd - 2
print("\nCheck control : number of lines read from Factos XLSX " + str(nblines) + "\n\nEverything is ok !")
