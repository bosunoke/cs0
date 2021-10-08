import os
from openpyxl import load_workbook

# Init custome variable
xlsxpath= '/home/adminlflp/Scripts/Email_Lflp/listing_memo_emails_lflp.xlsx'

# Get current path
crpth = (os.getcwd())

# Outpout path for csv Outlook
path_csv = crpth + '/outpout_lflp/csv/'

# Outpout path to the outpout VCF file
path_vcf = crpth + '/outpout_lflp/vcf/contacts_lflp.vcf'
vcf = open(path_vcf, "w")

# Variables init
lstcontact = {}
lstgroups = []

# Access path to the xlsx source file with 5 columns in this specific order :
# firstname, lastname, group, email, nk2_rc_lflp (label for Outlook)
wb = load_workbook(filename = xlsxpath, data_only=True, read_only=True)
ws = wb['nk2_rc_lflp']

print("\nExcel file reading in progress\n")

indexrow = 1
indextuple = 1

for row in ws.rows:
    if indexrow > 1:
        lstcontact['fn', indextuple] = ws["{}{}".format("A", indexrow)].value
        lstcontact['ln', indextuple] = ws["{}{}".format("B", indexrow)].value
        lstcontact['email', indextuple] = ws["{}{}".format("C", indexrow)].value
        lstcontact['group', indextuple] = str(ws["{}{}".format("D", indexrow)].value)
        if lstcontact['group', indextuple] != "None" and lstcontact['group', indextuple] not in lstgroups:
            lstgroups.append(lstcontact['group', indextuple])
        indextuple += 1
    indexrow += 1

# Create csv file for each group
for entry in lstgroups:
    pathfilecsvadm = path_csv + '_' + str(entry) + '.csv'
    if os.path.isfile(pathfilecsvadm):
        os.remove(pathfilecsvadm)
    csvadm = open(pathfilecsvadm,"w",encoding="cp1252")
    csvadm.write('"firstname","lastname","email"\n')
    csvadm.close()

# Read the tuple Dict for writing files
indextd = 1
while indextd < indextuple:
    if lstcontact['group', indextd] != "None":
        # Write vcf file
        vcf.write('BEGIN:VCARD\nVERSION:3.0\nN:' + lstcontact['fn', indextd] + ';' + lstcontact['ln', indextd] +
                  ';;;\nFN:' + lstcontact['fn', indextd] + " " + lstcontact['ln', indextd] +
                  '\nEMAIL;TYPE=INTERNET;TYPE=WORK:' + lstcontact['email', indextd] +
                  '\nCATEGORIES:' + lstcontact['group', indextd] + '\nEND:VCARD\n')
        # Write csv files
        pathfilecsvadm = pathfilecsvadm = path_csv + '_' + lstcontact['group', indextd] + '.csv'
        csvadm = open(pathfilecsvadm, "a",encoding="cp1252")
        csvadm.write('"' + lstcontact['fn', indextd] + '","' + lstcontact['ln', indextd] \
                     + '","' + lstcontact['email', indextd] + '"\n')
        csvadm.close()
    indextd += 1

vcf.close()
print("\n\nEverything is ok ! Check your new VCard file : " + path_vcf)
print("\nEverything is ok ! Check your new CSV file : " + path_csv)
